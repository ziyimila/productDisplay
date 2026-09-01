#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""从 Excel 内容模板生成 miniprogram/data/packages.js

用法:
    python tools/import_content.py                          # 默认读取 outputs/.../装修套餐内容模板.xlsx
    python tools/import_content.py 路径/模板.xlsx           # 指定模板
    python tools/import_content.py --out tmp/packages.js    # 输出到指定文件（不覆盖原文件）
    python tools/import_content.py --dry-run                # 只校验不写文件

依赖: openpyxl（项目自带运行时 Python 已包含）
"""
import argparse
import re
import sys
from collections import Counter
from pathlib import Path

ROOT = Path(__file__).resolve().parent.parent
DEFAULT_XLSX = ROOT / "outputs" / "01a05337-61ef-7090-9ab3-d4456627ff26" / "装修套餐内容模板.xlsx"
DEFAULT_OUT = ROOT / "miniprogram" / "data" / "packages.js"

SHEET_RULES = "计价规则"
SHEET_PACKAGES = "套餐"
SHEET_CATEGORIES = "分类"
SHEET_BRANDS = "品牌"
SHEET_MODELS = "型号"

QUERY_HELPERS = '''
/* ---------------- 查询工具 ---------------- */

function getPackageById(id) {
  return PACKAGES.find((p) => p.id === id)
}

function getCategoryById(id) {
  return CATEGORIES.find((c) => c.id === id)
}

function getBrandById(id) {
  return BRANDS.find((b) => b.id === id)
}

function getModelById(id) {
  return MODELS.find((m) => m.id === id)
}

function getModelsByCategory(categoryId) {
  return MODELS.filter((m) => m.categoryId === categoryId)
}

function getModelsByCategoryAndBrand(categoryId, brandId) {
  return MODELS.filter((m) => m.categoryId === categoryId && m.brandId === brandId)
}

// 某个分类下套餐标配（不加价）的所有型号
function getStandardModels(categoryId) {
  return getModelsByCategory(categoryId).filter((m) => m.type === 'standard')
}

// 默认标配型号（取第一款，客户可自由换成其他标配款）
function getStandardModel(categoryId) {
  const list = getStandardModels(categoryId)
  return list[0] || null
}

// 某个分类下有哪些品牌（按该分类型号出现的顺序去重）
function getBrandsOfCategory(categoryId) {
  const result = []
  getModelsByCategory(categoryId).forEach((m) => {
    if (!result.some((b) => b.id === m.brandId)) {
      const brand = getBrandById(m.brandId)
      if (brand) result.push(brand)
    }
  })
  return result
}

// 套餐完整信息：套餐 + 每个分类（含品牌及品牌下的型号）
function getPackageFull(packageId) {
  const pkg = getPackageById(packageId)
  if (!pkg) return null
  const categories = (pkg.categories || []).map((cid) => {
    const cat = getCategoryById(cid)
    if (!cat) return null
    const brands = getBrandsOfCategory(cid).map((b) => ({
      id: b.id,
      name: b.name,
      models: getModelsByCategoryAndBrand(cid, b.id)
    }))
    return Object.assign({}, cat, { brands })
  }).filter(Boolean)
  return Object.assign({}, pkg, { categories })
}

module.exports = {
  CATEGORIES,
  BRANDS,
  MODELS,
  PACKAGES,
  RULES,
  getPackageById,
  getCategoryById,
  getBrandById,
  getModelById,
  getModelsByCategory,
  getModelsByCategoryAndBrand,
  getStandardModels,
  getStandardModel,
  getBrandsOfCategory,
  getPackageFull
}
'''


def load_workbook(path):
    try:
        from openpyxl import load_workbook
    except ImportError:
        sys.exit("缺少 openpyxl，请先安装：pip install openpyxl （或使用项目自带运行时 Python）")
    return load_workbook(path, data_only=True)


def cell_str(value):
    if value is None:
        return ""
    return str(value).strip()


def to_number(value, field, row_label, errors, allow_blank=False):
    if value is None or cell_str(value) == "":
        if allow_blank:
            return 0
        errors.append(f"{field}（{row_label}）为空，需要填数字")
        return 0
    s = cell_str(value).replace(",", "").replace("，", "").replace("￥", "").replace("¥", "")
    try:
        n = float(s)
    except ValueError:
        errors.append(f"{field}（{row_label}）不是数字：{s!r}")
        return 0
    return int(n) if n.is_integer() else n


def parse_rules(sheet, errors, warnings):
    rules = {}
    for row in sheet.iter_rows(min_row=2):
        code = cell_str(row[0].value)
        if not code or not code.startswith(("stdArea", "stdRoom.", "extraRates.")):
            continue
        val = to_number(row[2].value, f"计价规则 {code}", code, errors, allow_blank=True)
        if code == "stdArea":
            rules["stdArea"] = val
        elif code.startswith("stdRoom."):
            rules.setdefault("stdRoom", {})[code.split(".", 1)[1]] = int(val)
    if "stdArea" not in rules:
        errors.append("计价规则表缺少 stdArea（标准覆盖面积）")
    return rules


def parse_categories(sheet):
    rows = []
    for row in sheet.iter_rows(min_row=2):
        cid = cell_str(row[0].value)
        if not cid:
            continue
        rows.append({"id": cid, "name": cell_str(row[1].value), "icon": cell_str(row[2].value)})
    return rows


def parse_brands(sheet):
    rows = []
    for row in sheet.iter_rows(min_row=2):
        bid = cell_str(row[0].value)
        if not bid:
            continue
        rows.append({"id": bid, "name": cell_str(row[1].value)})
    return rows


def _is_note_row(row):
    """说明/提示行：A 列含中文或过长，不是数据行"""
    a = cell_str(row[0].value)
    return bool(a) and (len(a) > 24 or any('\u4e00' <= c <= '\u9fff' for c in a))


def _split_mat_options(value):
    """把「品牌生态板（兔宝宝、大王椰、莫干山、千年舟）」拆成可选项列表 + 类型前缀"""
    m = re.match(r"^(.*?)（([^（）]+)）$", value)
    if m:
        prefix = m.group(1).strip()
        opts = [x.strip() for x in re.split(r"[、，,;；\s]+", m.group(2)) if x.strip()]
        if opts:
            return opts, prefix
    opts = [x.strip() for x in re.split(r"[、，,;；\s]+", value) if x.strip()]
    return opts, ""


def parse_packages(sheet, warnings):
    rows = []
    for row in sheet.iter_rows(min_row=2):
        pid = cell_str(row[0].value)
        if not pid:
            continue
        if _is_note_row(row):
            if any(cell_str(c.value) for c in row[1:]):
                warnings.append(f"「套餐」表第 {row[0].row} 行被跳过（疑似说明文字，请检查是否漏填数据）")
            continue
        hot = cell_str(row[5].value)
        categories = [c.strip() for c in re.split(r"[，,;；\n]", cell_str(row[6].value)) if c.strip()]
        # 业务规则：超出户型只有卫生间按「间」单独计价，其他超出房间按面积核算、不单独加价
        extra_rates = {"bathroom": row[7].value}
        materials = []
        for seg in re.split(r"[；;]", cell_str(row[8].value)):
            seg = seg.strip()
            if not seg:
                continue
            m = re.match(r"^(.+?)[：:](.*)$", seg)
            if m:
                label = m.group(1).strip()
                value = m.group(2).strip()
                opts, prefix = _split_mat_options(value)
                materials.append({"label": label, "value": value, "options": opts, "prefix": prefix})
            else:
                materials.append({"label": seg, "value": "", "options": [seg], "prefix": ""})
        rows.append({
            "id": pid,
            "name": cell_str(row[1].value),
            "basePrice": 0,  # placeholder, filled in validate step
            "perSqmRate": 0,
            "tag": cell_str(row[4].value),
            "hot": hot == "是",
            "categories": categories,
            "materials": materials,
            "_basePrice_raw": row[2].value,
            "_perSqmRate_raw": row[3].value,
            "_extraRates_raw": extra_rates,
        })
    return rows


def parse_models(sheet):
    rows = []
    for row in sheet.iter_rows(min_row=2):
        mid = cell_str(row[0].value)
        if not mid:
            continue
        specs = []
        specs_text = cell_str(row[8].value)
        for line in specs_text.splitlines():
            line = line.strip().rstrip("，,；;")
            if not line:
                continue
            m = re.match(r"^(.+?)[：:](.*)$", line)
            if m:
                specs.append({"k": m.group(1).strip(), "v": m.group(2).strip()})
            else:
                specs.append({"k": line, "v": ""})
        images = [i.strip() for i in re.split(r"[，,;；\n]", cell_str(row[7].value)) if i.strip()]
        rows.append({
            "id": mid,
            "categoryId": cell_str(row[1].value),
            "brandId": cell_str(row[2].value),
            "name": cell_str(row[3].value),
            "type": cell_str(row[4].value),
            "extraPrice": 0,
            "emoji": cell_str(row[6].value),
            "images": images,
            "specs": specs,
            "desc": cell_str(row[9].value),
            "_extraPrice_raw": row[5].value,
        })
    return rows


def validate(categories, brands, packages, models, errors, warnings):
    def dup_errors(items, label):
        c = Counter(i["id"] for i in items)
        for vid, n in c.items():
            if n > 1:
                errors.append(f"{label}表 id 重复：{vid}（出现 {n} 次）")

    dup_errors(categories, "分类")
    dup_errors(brands, "品牌")
    dup_errors(packages, "套餐")
    dup_errors(models, "型号")

    cat_ids = {c["id"] for c in categories}
    brand_ids = {b["id"] for b in brands}

    for p in packages:
        p["basePrice"] = to_number(p.pop("_basePrice_raw"), f"套餐 {p['id']} 一口价", p["id"], errors)
        p["perSqmRate"] = to_number(p.pop("_perSqmRate_raw"), f"套餐 {p['id']} 超面积单价", p["id"], errors)
        # 业务规则：超出户型只有卫生间单独计价，其他超出房间按面积核算、不单独加价
        raw = p.pop("_extraRates_raw", {})
        val = to_number(raw.get("bathroom"), f"套餐 {p['id']} 超出卫生间单价", p["id"], errors)
        if val < 0:
            errors.append(f"套餐 {p['id']} 超出卫生间单价不能为负数（当前 {val}）")
        p["extraRates"] = {"bathroom": int(val)}
        for cid in p["categories"]:
            if cid not in cat_ids:
                errors.append(f"套餐 {p['id']} 的包含分类 {cid!r} 在「分类」表中不存在")

    type_map = {"标配": "standard", "升级": "upgrade"}
    std_by_cat = Counter()
    all_cat_of_pkg = set()
    for p in packages:
        all_cat_of_pkg.update(p["categories"])
    for m in models:
        if m["type"] not in type_map:
            errors.append(f"型号 {m['id']} 的类型 {m['type']!r} 只能是「标配」或「升级」")
            continue
        m["type"] = type_map[m["type"]]
        m["extraPrice"] = to_number(m.pop("_extraPrice_raw"), f"型号 {m['id']} 升级补差价", m["id"], errors, allow_blank=True)
        if m["categoryId"] not in cat_ids:
            errors.append(f"型号 {m['id']} 的分类ID {m['categoryId']!r} 在「分类」表中不存在")
        if m["brandId"] not in brand_ids:
            errors.append(f"型号 {m['id']} 的品牌ID {m['brandId']!r} 在「品牌」表中不存在")
        if m["type"] == "standard":
            std_by_cat[m["categoryId"]] += 1
            if m["extraPrice"] != 0:
                errors.append(f"型号 {m['id']} 是标配但升级补差价不是 0（当前 {m['extraPrice']}）")
        for spec in m["specs"]:
            if not spec["k"] or not spec["v"]:
                warnings.append(f"型号 {m['id']} 的规格行缺少「项目：内容」格式：{spec['k'] or spec['v']}")

    for cid in sorted(all_cat_of_pkg):
        if std_by_cat[cid] < 1:
            errors.append(f"分类 {cid} 被套餐包含，但没有任何「标配」型号（客户将无法不加价选择）")
    for c in categories:
        if c["id"] not in all_cat_of_pkg:
            warnings.append(f"分类 {c['id']} 未被任何套餐包含（如果不需要可以删掉）")


def js_str(s):
    return "'" + s.replace("\\", "\\\\").replace("'", "\\'").replace("\n", "\\n").replace("\r", "") + "'"


def js_num(obj, key):
    v = obj.get(key, 0)
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def build_js(categories, brands, models, rules, packages):
    head = """/**
 * 套餐 / 分类 / 品牌 / 型号 数据
 * 本文件由 tools/import_content.py 从 Excel 内容模板自动生成，请勿直接手改；
 * 修改数据请编辑 Excel 模板，然后重新运行导入脚本。
 * type: standard = 套餐标配（不加价，可有多个不同品牌款式）；upgrade = 升级选配（extraPrice 为补差价）。
 */

"""
    cat_block = "const CATEGORIES = [\n" + "\n".join(
        f"  {{ id: {js_str(c['id'])}, name: {js_str(c['name'])}, icon: {js_str(c['icon'])} }}," for c in categories
    ) + "\n]\n"
    brand_block = "const BRANDS = [\n" + "\n".join(
        f"  {{ id: {js_str(b['id'])}, name: {js_str(b['name'])} }}," for b in brands
    ) + "\n]\n"

    model_block = "const MODELS = [\n"
    for m in models:
        lines = ["  {"]
        lines.append(
            f"    id: {js_str(m['id'])}, categoryId: {js_str(m['categoryId'])}, brandId: {js_str(m['brandId'])}, "
            f"name: {js_str(m['name'])}, type: {js_str(m['type'])}, extraPrice: {js_num(m, 'extraPrice')},"
        )
        emoji = m.get("emoji", "")
        if emoji:
            lines.append(f"    emoji: {js_str(emoji)},")
        images = m.get("images") or []
        lines.append(f"    images: [{', '.join(js_str(i) for i in images)}],")
        specs = m.get("specs") or []
        if specs:
            lines.append("    specs: [")
            for spec in specs:
                lines.append(f"      {{ k: {js_str(spec['k'])}, v: {js_str(spec['v'])} }},")
            lines.append("    ],")
        lines.append(f"    desc: {js_str(m.get('desc', ''))}")
        lines.append("  },")
        model_block += "\n".join(lines) + "\n"
    model_block += "]\n"

    sr = rules.get("stdRoom", {})
    rules_block = (
        "// 套餐计价规则（由 Excel「计价规则」表生成）\n"
        "// 套餐为一口价，覆盖标准面积与标准户型；超出标准面积按套餐单价/㎡核算，\n"
        "// 超出户型只有卫生间按该套餐 extraRates 单价（元/间）单独核算，其他超出房间按面积核算。\n"
        "const RULES = {\n"
        f"  stdArea: {js_num(rules, 'stdArea')},\n"
        "  stdRoom: {\n"
        + "\n".join(f"    {k}: {int(v)}," for k, v in sr.items())
        + "\n  }\n"
        "}\n"
    )

    pkg_block = "const PACKAGES = [\n"
    for p in packages:
        hot = "true" if p["hot"] else "false"
        rates = f"bathroom: {int(p['extraRates']['bathroom'])}"
        mats = p.get("materials") or []
        mats_js = ""
        if mats:
            inner = ", ".join(
                "{ label: %s, value: %s, options: [%s], prefix: %s }" % (
                    js_str(m["label"]), js_str(m["value"]),
                    ", ".join(js_str(o) for o in m.get("options") or []),
                    js_str(m.get("prefix") or ""),
                ) for m in mats
            )
            mats_js = ", materials: [" + inner + "]"
        pkg_block += (f"  {{ id: {js_str(p['id'])}, name: {js_str(p['name'])}, basePrice: {p['basePrice']}, "
                      f"perSqmRate: {p['perSqmRate']}, tag: {js_str(p['tag'])}, hot: {hot}, "
                      f"categories: [{', '.join(js_str(c) for c in p['categories'])}], extraRates: {{ {rates} }}{mats_js} }},\n")
    pkg_block += "]\n"

    return head + "\n" + cat_block + "\n" + brand_block + "\n" + model_block + "\n" + rules_block + "\n" + pkg_block + "\n" + QUERY_HELPERS


def main():
    if hasattr(sys.stdout, "reconfigure"):
        try:
            sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        except Exception:
            pass
    ap = argparse.ArgumentParser(description="把 Excel 内容模板导入为 miniprogram/data/packages.js")
    ap.add_argument("xlsx", nargs="?", default=str(DEFAULT_XLSX))
    ap.add_argument("--out", default=str(DEFAULT_OUT))
    ap.add_argument("--dry-run", action="store_true", help="只校验，不写文件")
    args = ap.parse_args()

    xlsx_path = Path(args.xlsx)
    out_path = Path(args.out)
    if not xlsx_path.exists():
        sys.exit(f"找不到模板文件：{xlsx_path}")

    wb = load_workbook(xlsx_path)
    missing = [s for s in (SHEET_RULES, SHEET_PACKAGES, SHEET_CATEGORIES, SHEET_BRANDS, SHEET_MODELS) if s not in wb.sheetnames]
    if missing:
        sys.exit(f"模板缺少工作表：{', '.join(missing)}")

    errors, warnings = [], []
    categories = parse_categories(wb[SHEET_CATEGORIES])
    brands = parse_brands(wb[SHEET_BRANDS])
    packages = parse_packages(wb[SHEET_PACKAGES], warnings)
    models = parse_models(wb[SHEET_MODELS])
    rules = parse_rules(wb[SHEET_RULES], errors, warnings)
    validate(categories, brands, packages, models, errors, warnings)

    for wmsg in warnings:
        print(f"[提示] {wmsg}")
    if errors:
        print("\n校验失败，共 {} 个错误：".format(len(errors)))
        for emsg in errors:
            print(f"  ✗ {emsg}")
        sys.exit(1)

    js = build_js(categories, brands, models, rules, packages)
    if args.dry_run:
        print(f"[OK] 校验通过：分类 {len(categories)}、品牌 {len(brands)}、型号 {len(models)}、套餐 {len(packages)}")
        return
    out_path.parent.mkdir(parents=True, exist_ok=True)
    out_path.write_text(js, encoding="utf-8")
    print(f"[OK] 已生成 {out_path}（分类 {len(categories)}、品牌 {len(brands)}、型号 {len(models)}、套餐 {len(packages)}）")


if __name__ == "__main__":
    main()
